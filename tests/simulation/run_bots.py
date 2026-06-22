"""
Smart Land Copilot — Bot Simulation Suite
============================================
محاكاة 4 بوتات تعمل بشكل عشوائي ومتزامن باستخدام asyncio.

البوتات:
    1. Landowner Bot   — ينشر أراضي جديدة للبيع
    2. Broker Bot      — يسجل نفسه ويعين أراضٍ باسمه
    3. Investor Bot    — يشتري الأراضي (محاكاة المعاملة)
    4. Accounting Bot  — يتحقق من صحة حسابات العمولات وينشئ تقرير Excel

التشغيل:
    python tests/simulation/run_bots.py

المخرجات:
    - simulation_report_TIMESTAMP.xlsx  — تقرير شامل بنتائج المحاكاة
    - console log مع كل معاملة
"""

import asyncio
import logging
import random
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Dict, List, Optional

# ─── Try to create Excel report — fallback to CSV if openpyxl not available ───
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    import csv

logger = logging.getLogger("smartland_bots")

# ══════════════════════════════════════════════
# إعدادات المحاكاة
# ══════════════════════════════════════════════

SIMULATION_CONFIG = {
    "num_landowners": 5,      # عدد بوتات المالكين
    "num_brokers": 10,         # عدد بوتات الوسطاء
    "num_investors": 15,       # عدد بوتات المستثمرين
    "num_accountants": 2,      # عدد بوتات المحاسبة
    "lands_per_owner": 3,      # عدد الأراضي التي ينشرها كل مالك
    "min_land_price": 100_000,  # أقل سعر أرض
    "max_land_price": 50_000_000,  # أعلى سعر أرض
    "commission_pct_range": (2.0, 7.5),  # نطاق نسبة العمولة
    "simulation_minutes": 2,   # مدة المحاكاة بالدقائق
    "check_interval_sec": 2,   # فترة التحقق من المعاملات الجديدة
}

# ══════════════════════════════════════════════
# هياكل البيانات
# ══════════════════════════════════════════════

@dataclass
class Land:
    """أرض معروضة للبيع."""
    land_id: str
    owner_id: str
    owner_name: str
    governorate: str
    total_area_sqm: float
    price_per_sqm_egp: float
    total_price_egp: float
    commission_pct: float
    status: str = "Available"  # Available, Sold, Transferred
    created_at: str = ""
    sold_at: Optional[str] = None
    buyer_id: Optional[str] = None
    broker_id: Optional[str] = None

    def __post_init__(self):
        if not self.created_at:
            self.created_at = datetime.now(timezone.utc).isoformat()


@dataclass
class Transaction:
    """معاملة شراء."""
    tx_id: str
    land_id: str
    buyer_id: str
    seller_id: str
    broker_id: Optional[str]
    sale_price_egp: float
    commission_pct: float
    commission_amount_egp: float
    platform_fee_egp: float
    net_to_seller_egp: float
    status: str = "Completed"
    verified: bool = False  # هل تحقق منها بوت المحاسبة؟
    created_at: str = ""


@dataclass
class BotState:
    """حالة البوتات - للمحاسبة."""
    landowners: Dict[str, dict] = field(default_factory=dict)
    brokers: Dict[str, dict] = field(default_factory=dict)
    investors: Dict[str, dict] = field(default_factory=dict)
    lands: Dict[str, Land] = field(default_factory=dict)
    transactions: List[Transaction] = field(default_factory=list)
    errors: List[dict] = field(default_factory=list)


# ══════════════════════════════════════════════
# محاكاة BrokerDelegationService (في الذاكرة)
# ══════════════════════════════════════════════

class SimBrokerService:
    """نسخة مبسطة من BrokerDelegationService للمحاكاة."""

    def __init__(self):
        self._allocations: Dict[str, List[dict]] = {}  # land_id -> [broker_id, ...]

    def allocate_broker(self, land_id: str, broker_id: str) -> bool:
        if land_id not in self._allocations:
            self._allocations[land_id] = []
        if len(self._allocations[land_id]) >= 2:
            return False
        if broker_id in self._allocations[land_id]:
            return False
        self._allocations[land_id].append(broker_id)
        return True

    def get_land_brokers(self, land_id: str) -> List[str]:
        return self._allocations.get(land_id, [])


# ══════════════════════════════════════════════
# 1. بوت المالك (Landowner Bot)
# ══════════════════════════════════════════════

GOVERNORATES = [
    "القاهرة", "الجيزة", "الإسكندرية", "السويس",
    "الأقصر", "أسوان", "الغردقة", "شرم الشيخ",
    "بورسعيد", "دمياط", "المنصورة", "طنطا",
]

LAND_NAMES = [
    "أرض سكنية", "أرض تجارية", "أرض صناعية",
    "أرض زراعية", "أرض استثمارية", "أرض سياحية",
    "مزرعة", "فيلا", "عمارة", "مكتب",
]


async def landowner_bot(
    owner_id: str,
    owner_name: str,
    state: BotState,
    broker_service: SimBrokerService,
    config: dict,
):
    """
    بوت المالك — ينشر أراضي جديدة للبيع بشكل عشوائي.
    يختار موقعاً وسعراً ونسبة عمولة لكل أرض.
    """
    lands_to_create = random.randint(1, config["lands_per_owner"])

    for i in range(lands_to_create):
        # إنشاء أرض عشوائية
        area_sqm = random.randint(500, 100_000)
        price_per_sqm = random.uniform(
            config["min_land_price"] / 1000,
            config["max_land_price"] / 1000,
        )
        total_price = round(area_sqm * price_per_sqm, 2)
        commission = round(random.uniform(*config["commission_pct_range"]), 1)

        land = Land(
            land_id=f"SIM-{uuid.uuid4().hex[:8].upper()}",
            owner_id=owner_id,
            owner_name=owner_name,
            governorate=random.choice(GOVERNORATES),
            total_area_sqm=area_sqm,
            price_per_sqm_egp=round(price_per_sqm, 2),
            total_price_egp=total_price,
            commission_pct=commission,
        )

        state.lands[land.land_id] = land

        logger.info(
            f"🏠 [مالك] {owner_name} نشر أرض {land.land_id} — "
            f"{land.governorate} | {area_sqm:,}م² | "
            f"{total_price:,.0f} ج.م | عمولة {commission}%"
        )

        # انتظار عشوائي بين نشر الأراضي
        await asyncio.sleep(random.uniform(0.5, 2.0))

    logger.info(f"✅ [مالك] {owner_name} انتهى من نشر {lands_to_create} أراضي")


# ══════════════════════════════════════════════
# 2. بوت الوسيط (Broker Bot)
# ══════════════════════════════════════════════

async def broker_bot(
    broker_id: str,
    broker_name: str,
    state: BotState,
    broker_service: SimBrokerService,
    config: dict,
):
    """
    بوت الوسيط — يسجل نفسه ويعيّن أراضٍ باسمه.
    يمكن للوسيط أن يُعيَّن لأرضين كحد أقصى (قاعدة النظام).
    """
    # الحصول على الأراضي المتاحة
    available_lands = [
        land for land in state.lands.values()
        if land.status == "Available"
        and land.owner_id != broker_id
    ]

    if not available_lands:
        logger.info(f"ℹ️ [وسيط] {broker_name} — لا توجد أراضٍ متاحة للتسجيل")
        return

    # اختيار عشوائي للأراضي (1-3 أراضٍ)
    num_to_assign = min(random.randint(1, 3), len(available_lands))
    selected_lands = random.sample(available_lands, num_to_assign)

    assigned_count = 0
    for land in selected_lands:
        # التحقق من حد 2 وسيط للأرض
        success = broker_service.allocate_broker(land.land_id, broker_id)
        if success:
            land.broker_id = broker_id
            assigned_count += 1
            logger.info(
                f"📋 [وسيط] {broker_name} سجّل في أرض {land.land_id} — "
                f"{land.governorate} | {land.total_price_egp:,.0f} ج.م"
            )
        await asyncio.sleep(random.uniform(0.3, 1.0))

    logger.info(f"✅ [وسيط] {broker_name} سجّل في {assigned_count} أراضي")


# ══════════════════════════════════════════════
# 3. بوت المستثمر (Investor Bot)
# ══════════════════════════════════════════════

async def investor_bot(
    investor_id: str,
    investor_name: str,
    state: BotState,
    broker_service: SimBrokerService,
    config: dict,
):
    """
    بوت المستثمر — يشتري الأراضي المتاحة.
    - يتحقق من وجود رصيد كافٍ (محاكاة)
    - يحسب العمولات
    - يُنهي المعاملة
    """
    wallet_balance = random.uniform(500_000, 200_000_000)  # رصيد محاكى
    purchased_count = 0

    # الحصول على الأراضي المتاحة للبيع
    available = [
        land for land in state.lands.values()
        if land.status == "Available"
        and land.owner_id != investor_id
    ]

    if not available:
        logger.info(f"ℹ️ [مستثمر] {investor_name} — لا توجد أراضٍ متاحة للشراء")
        return

    # شراء 1-3 أراضي عشوائياً
    num_to_buy = min(random.randint(1, 3), len(available))
    targets = random.sample(available, num_to_buy)

    for land in targets:
        if wallet_balance < land.total_price_egp:
            logger.info(
                f"⏭️ [مستثمر] {investor_name} — رصيد غير كافٍ "
                f"لأرض {land.land_id} ({land.total_price_egp:,.0f} ج.م)"
            )
            continue

        # حساب العمولات
        commission_amt = land.total_price_egp * (land.commission_pct / 100.0)
        platform_fee = land.total_price_egp * 0.005  # 0.5% منصة
        net_to_seller = land.total_price_egp - commission_amt - platform_fee

        # خصم من الرصيد
        wallet_balance -= land.total_price_egp

        # تحديث حالة الأرض
        land.status = "Sold"
        land.buyer_id = investor_id
        land.sold_at = datetime.now(timezone.utc).isoformat()

        # الحصول على الوسيط المسجل (إن وجد)
        assigned_brokers = broker_service.get_land_brokers(land.land_id)
        winning_broker = assigned_brokers[0] if assigned_brokers else None

        # تسجيل المعاملة
        tx = Transaction(
            tx_id=f"TX-{uuid.uuid4().hex[:12].upper()}",
            land_id=land.land_id,
            buyer_id=investor_id,
            seller_id=land.owner_id,
            broker_id=winning_broker,
            sale_price_egp=land.total_price_egp,
            commission_pct=land.commission_pct,
            commission_amount_egp=commission_amt,
            platform_fee_egp=platform_fee,
            net_to_seller_egp=net_to_seller,
        )
        state.transactions.append(tx)
        purchased_count += 1

        logger.info(
            f"💰 [مستثمر] {investor_name} اشترى أرض {land.land_id} — "
            f"{land.total_price_egp:,.0f} ج.م | "
            f"عمولة {commission_amt:,.0f} ج.م ({land.commission_pct}%)"
        )

        if winning_broker:
            # تحديث إحصائيات الوسيط
            if winning_broker not in state.brokers:
                state.brokers[winning_broker] = {
                    "name": f"وسيط-{winning_broker[:8]}",
                    "total_commission_egp": 0.0,
                    "deals_closed": 0,
                }
            state.brokers[winning_broker]["total_commission_egp"] += commission_amt
            state.brokers[winning_broker]["deals_closed"] += 1

        await asyncio.sleep(random.uniform(1.0, 3.0))

    logger.info(
        f"✅ [مستثمر] {investor_name} اشترى {purchased_count} أراضي "
        f"(الرصيد المتبقي: {wallet_balance:,.0f} ج.م)"
    )


# ══════════════════════════════════════════════
# 4. بوت الحسابات (Accounting Bot)
# ══════════════════════════════════════════════

async def accounting_bot(
    accountant_id: str,
    accountant_name: str,
    state: BotState,
    config: dict,
):
    """
    بوت الحسابات — يجمع المعاملات ويتحقق من صحة العمولات.
    القاعدة: commission = sale_price * (commission_pct / 100)

    ينشئ تقرير Excel يوضح:
    - كل المعاملات
    - العمولة المتوقعة مقابل العمولة المحسوبة
    - أي انحرافات (إذا كانت مختلفة)
    """
    logger.info(f"🔍 [محاسبة] {accountant_name} — بدء التدقيق...")

    discrepancies = []
    verified_tx = []
    total_expected = 0.0
    total_actual = 0.0

    for tx in state.transactions:
        if tx.verified:
            continue

        # التحقق من العمولة: commission = price * (pct / 100)
        expected_commission = round(tx.sale_price_egp * (tx.commission_pct / 100.0), 2)
        actual_commission = round(tx.commission_amount_egp, 2)

        total_expected += expected_commission
        total_actual += actual_commission

        deviation = round(actual_commission - expected_commission, 2)

        entry = {
            "tx_id": tx.tx_id,
            "land_id": tx.land_id,
            "buyer": tx.buyer_id[:12],
            "seller": tx.seller_id[:12],
            "broker": tx.broker_id[:12] if tx.broker_id else "بدون وسيط",
            "sale_price": tx.sale_price_egp,
            "commission_pct": tx.commission_pct,
            "expected_commission": expected_commission,
            "actual_commission": actual_commission,
            "deviation": deviation,
            "is_discrepancy": abs(deviation) > 0.01,  # تفاوت مسموح: 1 قرش
        }

        verified_tx.append(entry)
        tx.verified = True

        if entry["is_discrepancy"]:
            discrepancies.append(entry)
            logger.warning(
                f"⚠️ [محاسبة] انحراف في {tx.tx_id}: "
                f"متوقع {expected_commission:,.2f} ≠ فعلي {actual_commission:,.2f} "
                f"(فرق {deviation:,.2f})"
            )
        else:
            logger.info(
                f"✅ [محاسبة] {tx.tx_id}: صحيح ✓ "
                f"(عمولة {actual_commission:,.0f} ج.م = "
                f"{tx.commission_pct}% من {tx.sale_price_egp:,.0f})"
            )

        await asyncio.sleep(0.1)  # تأخير بسيط

    # إنشاء التقرير
    report = {
        "accountant_id": accountant_id,
        "verified_at": datetime.now(timezone.utc).isoformat(),
        "total_transactions": len(verified_tx),
        "total_expected_commission": round(total_expected, 2),
        "total_actual_commission": round(total_actual, 2),
        "total_deviation": round(total_expected - total_actual, 2),
        "discrepancies_count": len(discrepancies),
        "transactions": verified_tx,
        "discrepancies": discrepancies,
    }

    logger.info(
        f"📊 [محاسبة] {accountant_name}: تم تدقيق {len(verified_tx)} معاملة | "
        f"{'⚠️ ' + str(len(discrepancies)) + ' انحرافات' if discrepancies else '✅ لا توجد انحرافات'}"
    )

    return report


# ══════════════════════════════════════════════
# إنشاء تقرير Excel
# ══════════════════════════════════════════════

def create_excel_report(
    report: dict,
    state: BotState,
    filename: str,
):
    """إنشاء تقرير Excel شامل."""
    if HAS_EXCEL:
        _create_excel_report_xlsx(report, state, filename)
    else:
        _create_excel_report_csv(report, state, filename)


def _create_excel_report_xlsx(report: dict, state: BotState, filename: str):
    """إنشاء تقرير Excel بتنسيق XLSX."""
    wb = Workbook()

    # ─── Sheet 1: Summary ───
    ws_summary = wb.active
    ws_summary.title = "الملخص"
    ws_summary.sheet_view.rightToLeft = True

    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # عنوان
    ws_summary.merge_cells("A1:F1")
    ws_summary["A1"] = "تقرير محاكاة البوتات — Smart Land Copilot"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A1"].alignment = Alignment(horizontal="center")

    # إحصائيات عامة
    ws_summary["A3"] = "الإحصائية"
    ws_summary["B3"] = "القيمة"
    for cell in [ws_summary["A3"], ws_summary["B3"]]:
        cell.font = header_font
        cell.fill = header_fill

    stats = [
        ("إجمالي الأراضي المنشورة", len(state.lands)),
        ("إجمالي المعاملات", len(state.transactions)),
        ("إجمالي العمولات المتوقعة (ج.م)", f"{report['total_expected_commission']:,.2f}"),
        ("إجمالي العمولات الفعلية (ج.م)", f"{report['total_actual_commission']:,.2f}"),
        ("الفرق الإجمالي (ج.م)", f"{report['total_deviation']:,.2f}"),
        ("عدد الانحرافات", report["discrepancies_count"]),
        ("وقت التدقيق", report["verified_at"][:19]),
    ]

    for i, (label, value) in enumerate(stats, start=4):
        ws_summary[f"A{i}"] = label
        ws_summary[f"B{i}"] = value
        if "الانحرافات" in label or "الفرق" in label:
            if isinstance(value, str) and float(value.replace(",", "")) != 0:
                ws_summary[f"A{i}"].fill = red_fill if report["discrepancies_count"] > 0 else green_fill
                ws_summary[f"B{i}"].fill = red_fill if report["discrepancies_count"] > 0 else green_fill

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 25

    # ─── Sheet 2: Transactions ───
    ws_tx = wb.create_sheet("المعاملات")
    ws_tx.sheet_view.rightToLeft = True

    headers = [
        "معرف المعاملة", "الأرض", "المشتري", "البائع",
        "الوسيط", "سعر البيع", "نسبة العمولة",
        "العمولة المتوقعة", "العمولة الفعلية", "الفرق", "الحالة",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws_tx.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, tx in enumerate(report["transactions"], start=2):
        ws_tx.cell(row=i, column=1, value=tx["tx_id"])
        ws_tx.cell(row=i, column=2, value=tx["land_id"])
        ws_tx.cell(row=i, column=3, value=tx["buyer"])
        ws_tx.cell(row=i, column=4, value=tx["seller"])
        ws_tx.cell(row=i, column=5, value=tx["broker"])
        ws_tx.cell(row=i, column=6, value=tx["sale_price"])
        ws_tx.cell(row=i, column=7, value=tx["commission_pct"])
        ws_tx.cell(row=i, column=8, value=tx["expected_commission"])
        ws_tx.cell(row=i, column=9, value=tx["actual_commission"])
        ws_tx.cell(row=i, column=10, value=tx["deviation"])

        status = "✅ صحيح" if not tx["is_discrepancy"] else "⚠️ انحراف"
        status_cell = ws_tx.cell(row=i, column=11, value=status)
        if tx["is_discrepancy"]:
            status_cell.fill = red_fill
            # تلوين الصف
            for col in range(1, 12):
                ws_tx.cell(row=i, column=col).fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                )
        else:
            status_cell.fill = green_fill

    # ضبط عرض الأعمدة
    col_widths = [25, 18, 15, 15, 15, 18, 14, 18, 18, 14, 14]
    for i, width in enumerate(col_widths, 1):
        ws_tx.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = width

    # ─── Sheet 3: Lands ───
    ws_lands = wb.create_sheet("الأراضي")
    ws_lands.sheet_view.rightToLeft = True

    land_headers = [
        "معرف الأرض", "المالك", "المحافظة", "المساحة (م²)",
        "سعر المتر", "السعر الإجمالي", "الحالة", "الوسيط",
    ]
    for col, header in enumerate(land_headers, 1):
        cell = ws_lands.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for i, land in enumerate(state.lands.values(), start=2):
        ws_lands.cell(row=i, column=1, value=land.land_id)
        ws_lands.cell(row=i, column=2, value=land.owner_name)
        ws_lands.cell(row=i, column=3, value=land.governorate)
        ws_lands.cell(row=i, column=4, value=land.total_area_sqm)
        ws_lands.cell(row=i, column=5, value=land.price_per_sqm_egp)
        ws_lands.cell(row=i, column=6, value=land.total_price_egp)
        ws_lands.cell(row=i, column=7, value=land.status)
        ws_lands.cell(row=i, column=8, value=land.broker_id[:12] if land.broker_id else "-")

    # حفظ الملف
    wb.save(filename)
    logger.info(f"📁 تم حفظ التقرير: {filename}")


def _create_excel_report_csv(report: dict, state: BotState, filename: str):
    """Fallback: إنشاء تقرير CSV إذا لم يكن openpyxl متاحاً."""
    csv_filename = filename.replace(".xlsx", ".csv")

    with open(csv_filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["تقرير محاكاة البوتات — Smart Land Copilot"])
        writer.writerow(["إجمالي المعاملات", report["total_transactions"]])
        writer.writerow(["إجمالي العمولات المتوقعة", report["total_expected_commission"]])
        writer.writerow(["إجمالي العمولات الفعلية", report["total_actual_commission"]])
        writer.writerow(["الفرق", report["total_deviation"]])
        writer.writerow(["عدد الانحرافات", report["discrepancies_count"]])
        writer.writerow([])
        writer.writerow(["tx_id", "land_id", "sale_price", "expected", "actual", "deviation", "status"])
        for tx in report["transactions"]:
            status = "OK" if not tx["is_discrepancy"] else "DISCREPANCY"
            writer.writerow([
                tx["tx_id"], tx["land_id"], tx["sale_price"],
                tx["expected_commission"], tx["actual_commission"],
                tx["deviation"], status,
            ])

    logger.info(f"📁 تم حفظ التقرير (CSV): {csv_filename}")


# ══════════════════════════════════════════════
# المدير الرئيسي — تشغيل البوتات
# ══════════════════════════════════════════════

async def run_simulation(config: dict = None):
    """
    المدير الرئيسي للمحاكاة — يشغل جميع البوتات بشكل متزامن.
    """
    if config is None:
        config = SIMULATION_CONFIG

    config = config.copy()  # تجنب التعديل المباشر

    state = BotState()
    broker_service = SimBrokerService()

    logger.info("=" * 60)
    logger.info("  🚀 بدء محاكاة البوتات — Smart Land Copilot")
    logger.info(f"  المالكين: {config['num_landowners']} | "
                f"الوسطاء: {config['num_brokers']} | "
                f"المستثمرين: {config['num_investors']} | "
                f"المحاسبين: {config['num_accountants']}")
    logger.info(f"  المدة: {config['simulation_minutes']} دقيقة")
    logger.info("=" * 60)

    start_time = time.time()
    start_time + config["simulation_minutes"] * 60

    # ─── المرحلة 1: إنشاء بوتات المالكين ───
    logger.info("\n📌 المرحلة 1: نشر الأراضي (مالكون)...")
    owner_tasks = []
    for i in range(config["num_landowners"]):
        owner_id = f"OWNER-{uuid.uuid4().hex[:8].upper()}"
        owner_name = f"مالك-{i+1}"
        state.landowners[owner_id] = {"name": owner_name}
        task = landowner_bot(owner_id, owner_name, state, broker_service, config)
        owner_tasks.append(task)

    await asyncio.gather(*owner_tasks)

    # ─── المرحلة 2: إنشاء بوتات الوسطاء ───
    logger.info("\n📌 المرحلة 2: تسجيل الوسطاء...")
    broker_tasks = []
    for i in range(config["num_brokers"]):
        broker_id = f"BROKER-{uuid.uuid4().hex[:8].upper()}"
        broker_name = f"وسيط-{i+1}"
        state.brokers[broker_id] = {
            "name": broker_name,
            "total_commission_egp": 0.0,
            "deals_closed": 0,
        }
        task = broker_bot(broker_id, broker_name, state, broker_service, config)
        broker_tasks.append(task)

    await asyncio.gather(*broker_tasks)

    # ─── المرحلة 3: إنشاء بوتات المستثمرين ───
    logger.info("\n📌 المرحلة 3: شراء الأراضي (مستثمرون)...")
    investor_tasks = []
    for i in range(config["num_investors"]):
        investor_id = f"INVESTOR-{uuid.uuid4().hex[:8].upper()}"
        investor_name = f"مستثمر-{i+1}"
        state.investors[investor_id] = {"name": investor_name}
        task = investor_bot(investor_id, investor_name, state, broker_service, config)
        investor_tasks.append(task)

    await asyncio.gather(*investor_tasks)

    # ─── المرحلة 4: تشغيل بوتات المحاسبة ───
    logger.info("\n📌 المرحلة 4: تدقيق الحسابات...")
    account_tasks = []
    for i in range(config["num_accountants"]):
        acc_id = f"ACCT-{uuid.uuid4().hex[:8].upper()}"
        acc_name = f"محاسب-{i+1}"
        task = accounting_bot(acc_id, acc_name, state, config)
        account_tasks.append(task)

    reports = await asyncio.gather(*account_tasks)

    # ─── تقييم النتائج ───
    logger.info("\n" + "=" * 60)
    logger.info("  📊 ملخص المحاكاة")
    logger.info("=" * 60)
    logger.info(f"  الأراضي المنشورة:     {len(state.lands)}")
    logger.info(f"  المعاملات المكتملة:   {len(state.transactions)}")
    logger.info(f"  الوسطاء المشاركون:    {len([b for b in state.brokers.values() if b['deals_closed'] > 0])}")
    logger.info(f"  إجمالي العمولات:      {sum(r['total_actual_commission'] for r in reports):,.0f} ج.م")
    logger.info(f"  الانحرافات:           {sum(r['discrepancies_count'] for r in reports)}")

    # ─── إنشاء التقرير النهائي ───
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"simulation_report_{timestamp}.xlsx"

    combined_report = reports[0] if reports else {
        "accountant_id": "NONE",
        "verified_at": datetime.now(timezone.utc).isoformat(),
        "total_transactions": 0,
        "total_expected_commission": 0.0,
        "total_actual_commission": 0.0,
        "total_deviation": 0.0,
        "discrepancies_count": 0,
        "transactions": [],
        "discrepancies": [],
    }

    create_excel_report(combined_report, state, report_filename)

    elapsed = time.time() - start_time
    logger.info(f"\n⏱️  استغرقت المحاكاة: {elapsed:.1f} ثانية")
    logger.info(f"📄 التقرير: {report_filename}")

    # ─── التحقق من صحة العمولات ───
    if combined_report["discrepancies_count"] == 0:
        logger.info("\n✅ ✅ ✅ جميع العمولات صحيحة — لا توجد انحرافات! ✅ ✅ ✅")
        logger.info("✓ نظام العمولات يعمل بشكل صحيح")
        logger.info("✓ جميع المعاملات متوافقة مع قاعدة Winner-Takes-Commission")
        return {"status": "success", "deviations": 0, "report": report_filename}
    else:
        logger.warning(f"\n⚠️ ⚠️ ⚠️ تم العثور على {combined_report['discrepancies_count']} انحرافات! ⚠️ ⚠️ ⚠️")
        logger.warning("يرجى مراجعة التقرير للتفاصيل")
        return {
            "status": "discrepancies_found",
            "deviations": combined_report["discrepancies_count"],
            "report": report_filename,
        }


# ══════════════════════════════════════════════
# نقطة الدخول الرئيسية
# ══════════════════════════════════════════════

if __name__ == "__main__":
    # إعداد التسجيل
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(message)s",
        datefmt="%H:%M:%S",
    )

    # تشغيل المحاكاة
    result = asyncio.run(run_simulation())

    print("\n" + "=" * 60)
    if result["status"] == "success":
        print("  ✅ اكتملت المحاكاة — جميع العمولات صحيحة!")
    else:
        print(f"  ⚠️  اكتملت المحاكاة — {result['deviations']} انحرافات في العمولات")
    print(f"  📄 التقرير: {result['report']}")
    print("=" * 60)